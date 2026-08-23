import datetime as dt
import re
from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

import openpyxl

_JST = ZoneInfo("Asia/Tokyo")

_OFF_MARKERS = {"", "○", "◎", "休", "-", "ー", "off", "OFF", "×", "公休", "希望休", "有給", "欠勤"}
_YEAR_MONTH_RE = re.compile(r"(\d{4})\s*年\s*(\d{1,2})\s*月")
_MAX_ROWS = 400
_MAX_HEADER_SEARCH_ROWS = 50
_MIN_STRING_CELLS = 3

# 氏名が複数行(6行)にまたがる結合セルになっているブロック内で、
# シフト記号の行(offset 0)から数えて出勤時刻・退勤時刻が入っている行のオフセット。
# 対応表のテンプレート(月間オペレーション確認表)を実データで検証して確認した値。
_START_TIME_ROW_OFFSET = 2
_END_TIME_ROW_OFFSET = 3


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _extract_time_hhmm(value) -> str | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, dt.datetime):
        t = value.time()
        return f"{t.hour:02d}:{t.minute:02d}"
    if isinstance(value, dt.time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, dt.timedelta):
        # "[h]:mm"のような経過時間形式のセルはopenpyxlでtimedeltaとして読み込まれる
        total_minutes = int(value.total_seconds() // 60) % (24 * 60)
        return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
    if isinstance(value, (int, float)):
        frac = float(value) % 1
        total_minutes = round(frac * 24 * 60) % (24 * 60)
        return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
    return None


def _looks_numeric(text: str) -> bool:
    try:
        float(text)
        return True
    except ValueError:
        return False


def _normalize_name(name: str) -> str:
    return re.sub(r"\s+", "", name)


def _names_match(table_name: str, user_name: str) -> bool:
    a = _normalize_name(table_name)
    b = _normalize_name(user_name)
    if not a or not b:
        return False
    return a == b or a in b or b in a


def _merge_range_at(ws, row: int, col: int):
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return merged_range
    return None


def _find_year_month(ws, today: dt.date) -> tuple[int, int]:
    max_row = min(ws.max_row or 0, 30)
    max_col = min(ws.max_column or 0, 20)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            match = _YEAR_MONTH_RE.search(_cell_text(cell.value))
            if match:
                return int(match.group(1)), int(match.group(2))
    return today.year, today.month


def _find_day_header(ws, year_month_hint: tuple[int, int]) -> dict[dt.date, int] | None:
    """月間シフト表のヘッダー行(日付が横に並ぶ行)を探す。
    セルが実際の日付(datetime)の場合と、1〜31の素の数値の場合の両方に対応する。"""
    max_row = min(ws.max_row or 0, _MAX_HEADER_SEARCH_ROWS)
    for row_idx in range(1, max_row + 1):
        cells = ws[row_idx]

        date_to_col: dict[dt.date, int] = {}
        prev_date = None
        date_ordered_ok = True
        for cell in cells:
            value = cell.value
            if isinstance(value, dt.datetime):
                date_value = value.date()
            elif isinstance(value, dt.date):
                date_value = value
            else:
                continue
            if prev_date is not None and (date_value - prev_date).days != 1:
                date_ordered_ok = False
            date_to_col[date_value] = cell.column
            prev_date = date_value
        if date_ordered_ok and len(date_to_col) >= 28:
            return date_to_col

        day_to_col: dict[int, int] = {}
        prev_day = 0
        day_ordered_ok = True
        for cell in cells:
            value = cell.value
            if isinstance(value, bool) or not isinstance(value, (int, float)):
                continue
            if float(value).is_integer() and 1 <= int(value) <= 31:
                day = int(value)
                if day <= prev_day:
                    day_ordered_ok = False
                day_to_col[day] = cell.column
                prev_day = day
        if day_ordered_ok and len(day_to_col) >= 28 and 1 in day_to_col:
            year, month = year_month_hint
            converted: dict[dt.date, int] = {}
            for day, col in day_to_col.items():
                try:
                    converted[dt.date(year, month, day)] = col
                except ValueError:
                    continue
            if converted:
                return converted

    return None


def _find_name_and_block(ws, row_idx: int, first_data_col: int):
    """データ列より左側から氏名を探す。結合セル(複数行にまたがる)のテキストを優先し、
    社員番号のような数値だけのセルは氏名の候補から除外する。
    見つかった場合は (氏名, ブロック開始行, ブロック終了行) を返す。"""
    merged_candidates: list[tuple[int, str, int, int]] = []
    plain_candidates: list[tuple[int, str, int, int]] = []
    seen_ranges = set()

    for col in range(first_data_col - 1, 0, -1):
        merged_range = _merge_range_at(ws, row_idx, col)
        if merged_range is not None:
            key = (merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col)
            if key in seen_ranges:
                continue
            seen_ranges.add(key)
            value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            text = _cell_text(value)
            if not text or _looks_numeric(text):
                continue
            height = merged_range.max_row - merged_range.min_row + 1
            if height >= 2:
                merged_candidates.append((col, text, merged_range.min_row, merged_range.max_row))
            else:
                plain_candidates.append((col, text, row_idx, row_idx))
        else:
            text = _cell_text(ws.cell(row=row_idx, column=col).value)
            if text and not _looks_numeric(text):
                plain_candidates.append((col, text, row_idx, row_idx))

    if merged_candidates:
        merged_candidates.sort(key=lambda c: -c[0])
        _, name, block_min, block_max = merged_candidates[0]
        return name, block_min, block_max
    if plain_candidates:
        plain_candidates.sort(key=lambda c: -c[0])
        _, name, block_min, block_max = plain_candidates[0]
        return name, block_min, block_max
    return None


def _effective_max_row(ws) -> int:
    return min(ws.max_row or 0, _MAX_ROWS)


def _meiten_columns(ws, date_to_col: dict[dt.date, int], max_row: int) -> set[int]:
    """「銘」担当が(誰か1人でも)いる日付の列を1シートにつき1回だけ調べる。
    従業員ごと・日付ごとに毎回列を再スキャンしていた旧実装より大幅に速い。"""
    remaining = set(date_to_col.values())
    found: set[int] = set()
    for row_idx in range(1, max_row + 1):
        if not remaining:
            break
        for col in list(remaining):
            value = ws.cell(row=row_idx, column=col).value
            if isinstance(value, str) and "銘" in value:
                found.add(col)
                remaining.discard(col)
    return found


def _build_entries_for_row(
    ws, row_idx: int, block_max: int, date_to_col: dict[dt.date, int], meiten_cols: set[int]
):
    """あるシフト記号行(row_idx)について、日付ごとの出勤エントリと休み確定日を組み立てる。
    出退勤時刻のように数値(時刻シリアル値)しか入らない行は自然に除外される。
    出勤/休みの判定は出退勤時刻の有無を優先する(シフト記号が「○」等でも
    実際の時刻が入っていれば出勤扱い。夕方出勤スタッフなど特定の役割を持たない場合に対応)。"""
    start_row = row_idx + _START_TIME_ROW_OFFSET
    end_row = row_idx + _END_TIME_ROW_OFFSET

    entries = []
    off_dates: set[dt.date] = set()
    for date_value, col in date_to_col.items():
        code_text = _cell_text(ws.cell(row=row_idx, column=col).value)
        start = _extract_time_hhmm(ws.cell(row=start_row, column=col).value) if start_row <= block_max else None
        end = _extract_time_hhmm(ws.cell(row=end_row, column=col).value) if end_row <= block_max else None

        if not (start and end):
            # 出退勤時刻が入っていない日は、シフト記号が明示的な休み記号のときだけ「休み確定」とする
            if code_text and code_text in _OFF_MARKERS:
                off_dates.add(date_value)
            continue

        # 出退勤時刻が入っている=出勤日。シフト記号が「○」「◎」等の場合は
        # 夕方出勤スタッフなどで特定の役割を持たないケースなので、役割は「遅」(遅番)扱いにする。
        role = code_text if code_text and code_text not in _OFF_MARKERS else "遅"
        add_meiten = "洋" in role and col not in meiten_cols
        entries.append(
            {
                "date": date_value.isoformat(),
                "start": start,
                "end": end,
                "role": role,
                "add_meiten": add_meiten,
            }
        )

    entries.sort(key=lambda e: e["date"])
    return entries, sorted(d.isoformat() for d in off_dates)


def _iter_employee_rows(ws, date_to_col: dict[dt.date, int]):
    """シートを走査し、日付列に文字(シフト記号)が並ぶ行を氏名ブロックの手がかりとして
    (氏名, row_idx, block_min, block_max) を1従業員につき1回だけ返す。"""
    first_data_col = min(date_to_col.values())
    max_row = _effective_max_row(ws)
    seen_names: set[str] = set()

    for row_idx in range(1, max_row + 1):
        string_cells = 0
        for col in date_to_col.values():
            value = ws.cell(row=row_idx, column=col).value
            if isinstance(value, str) and value.strip():
                string_cells += 1
        if string_cells < _MIN_STRING_CELLS:
            continue

        found = _find_name_and_block(ws, row_idx, first_data_col)
        if not found:
            continue
        name, block_min, block_max = found
        if name in seen_names:
            continue
        seen_names.add(name)
        yield name, row_idx, block_min, block_max


def _extract_shifts_for_name(ws, date_to_col: dict[dt.date, int], user_name: str):
    max_row = _effective_max_row(ws)
    meiten_cols = _meiten_columns(ws, date_to_col, max_row)
    for name, row_idx, block_min, block_max in _iter_employee_rows(ws, date_to_col):
        if not _names_match(name, user_name):
            continue
        entries, off_dates = _build_entries_for_row(ws, row_idx, block_max, date_to_col, meiten_cols)
        return name, entries, off_dates

    return None, [], []


_PREFERRED_SHEET_NAMES = ("配布用",)


def _ordered_worksheets(workbook):
    """「配布用」のような配布・印刷用シートがあれば、それを他のシートより優先して調べる。
    社内マスター用シート(OP確認表など)は列構成が異なることが多く、誤って先にマッチしてしまうのを防ぐ。"""
    preferred = [ws for ws in workbook.worksheets if ws.title in _PREFERRED_SHEET_NAMES]
    rest = [ws for ws in workbook.worksheets if ws.title not in _PREFERRED_SHEET_NAMES]
    return preferred + rest


def extract_shift_dates_from_excel(file_bytes: bytes, user_name: str) -> dict:
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    today = datetime.now(_JST).date()
    any_header_found = False

    for ws in _ordered_worksheets(workbook):
        year_month_hint = _find_year_month(ws, today)
        date_to_col = _find_day_header(ws, year_month_hint)
        if not date_to_col:
            continue
        any_header_found = True

        matched_name, shifts, off_dates = _extract_shifts_for_name(ws, date_to_col, user_name)
        if matched_name and (shifts or off_dates):
            return {"shifts": shifts, "off_dates": off_dates, "matched_name_in_table": matched_name, "note": ""}

    if any_header_found:
        note = "日付が並んだ表は見つかりましたが、登録名と一致する行が見つかりませんでした。登録名がシフト表内の表記と一致しているか確認してください。"
    else:
        note = "対応している表の形式(日付が横に並んだヘッダー行を含む月間シフト表)が見つかりませんでした。"

    return {"shifts": [], "off_dates": [], "matched_name_in_table": None, "note": note}


def extract_all_shifts_from_excel(file_bytes: bytes) -> dict[str, dict]:
    """ファイル内の全従業員の出勤日・休み確定日を抽出する。
    戻り値は {氏名: {"shifts": [...], "off_dates": [...]}} の形。"""
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    today = datetime.now(_JST).date()

    for ws in _ordered_worksheets(workbook):
        year_month_hint = _find_year_month(ws, today)
        date_to_col = _find_day_header(ws, year_month_hint)
        if not date_to_col:
            continue

        max_row = _effective_max_row(ws)
        meiten_cols = _meiten_columns(ws, date_to_col, max_row)
        results: dict[str, dict] = {}
        for name, row_idx, block_min, block_max in _iter_employee_rows(ws, date_to_col):
            entries, off_dates = _build_entries_for_row(ws, row_idx, block_max, date_to_col, meiten_cols)
            if entries or off_dates:
                results[name] = {"shifts": entries, "off_dates": off_dates}

        if results:
            return results

    return {}
